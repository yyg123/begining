#coding=utf-8
import csv
import xlsxwriter
import xlrd
import xlwt
import re
import openpyxl

def csv_to_xlsx():                        #.csv转.xlsx文件
    with open('1.csv', 'r' ) as f:
        read = csv.reader(f)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('data')  # 创建一个sheet表格
        l = 0
        for line in read:
            print(line)
            r = 0
            for i in line:
                print(i)
                sheet.write(l, r, i)  # 一个一个将单元格数据写入
                r = r + 1
            l = l + 1

        workbook.save('1.xlsx')  # 保存Excel
if __name__ == '__main__':
    csv_to_xlsx()


# worksheet = xldr.open_workbook('1.xlsx')
# sheet_names = worksheet.sheet_names()
# print(sheet_names)
path = r'C:\Users\Administrator\Desktop\excel'
file = '1.xlsx'

data = xlrd.open_workbook(path + '/' + file)   #打开文件
nums = len(data.sheets())        #获取表格数目
for i in range(nums):            #打开sheet
    sheet1 = data.sheet()[i]

#sheet2 = data.sheet_by_name('1.xlsx')

nrows = sheet1.nrows         #行
ncols = sheet1.ncols         #列
print(nrows,ncols)

for i in range(nrows):         #循环行列表数据
    print(sheet1.row_values(i))

cell_A1 = sheet1.cell(0,0).value #获取单元格数据
print(cell_A1)

cell2_A2 = sheet2.row(0)[1].value
print(cell_A2)

#########################################################################################
def Exceldivide(file_dir):
 wb=openpyxl.load_workbook(file_dir)         #打开原有的excel表
 sheet=wb.get_sheet_by_name('Sheet1')
 tuple(sheet['A1':'C3'])

 wb.create_sheet('Sheet2')                 #新建一个表
 sheet2=wb.get_sheet_by_name('Sheet2')
 tuple(sheet2['A1':'C3'])

 L1=re.compile(r'\d\d/\d\d/\d\d\d\d')      #日期格式
 L2=re.compile(r'[a-zA-Z0-9_]+@[a-zA-Z0-9-]+.com')   #邮件格式
 l1=[]
 l2=[]
 for rows in sheet['A1':'C3']:           #提取日期和邮件数据
     for cell in rows:
         A=L1.search(cell.value)
         a=A.group()
         B=L2.search(cell.value)
         b=B.group()
 for rows in sheet2['A1':'A9']:         #把日期数据写入新表
    for cell in rows:
        cell.value=a
        print(cell.coordinate,cell.value)
 for rows in sheet2['B1':'B9']:        #把邮件数据写入新表
    for cell in rows:
        cell.value=b
        print(cell.coordinate,cell.value)
 return wb

g=Exceldivide('C:\\Users\\Desktop\\111_copy.xlsx')
g.save('C:\\Users\\Desktop\\111_copy.xlsx')    #保存
###################################################################################
import openpyxl
wb=openpyxl.load_workbook('ttt.xlsx')  #打开excel文件
print(wb.get_sheet_names())  #获取工作簿所有工作表名

sheet=wb.get_sheet_by_name('Sheet1')  #获取工作表
print(sheet.title)

sheet02=wb.get_active_sheet()  #获取活动的工作表
print(sheet02.title)


print(sheet['A1'].value)  #获取单元格A1值
print(sheet['A1'].column)  #获取单元格列值
print(sheet['A1'].row)  #获取单元格行号

print(sheet.cell(row=1,column=1).value)  #获取单元格A1值，column与row依然可用

for i in range(1,4,1):
    print(sheet.cell(row=i,column=1).value) #更加方便实用

print(sheet.max_column)  #获取最大列数
print(sheet.max_row)  #获取最大行数


#读取excel文件 sheetname可为空
def readwb(wbname,sheetname):
    wb=openpyxl.load_workbook(filename=wbname,read_only=True)
    if (sheetname==""):
        ws=wb.active
    else:
        ws=wb[sheetname]
    i=1
    fields=[]
    data=[]
    for row in ws.rows:
        list=[]
        for cell in row:
            aa=str(cell.value)
            if (aa==""):
                aa="1"
            list.append(aa)
        if(i<5):
            pass
        elif (i==5):
            fields=list
        else:
            data.append(list)
        i=i+1
    data.sort(key=lambda x:x[0])
    print (wbname +"-"+sheetname+"- 已成功读取")
    return fields,data


#新建excel
def creatwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

# 写入excel文件中 date 数据， fields 表头
def savetoexcel(data,fields,sheetname,wbname):
    print("写入excel：")
    wb=openpyxl.load_workbook(filename=wbname)

    sheet=wb.active
    sheet.title=sheetname

    field=1
    for field in range(1,len(fields)+1):   # 写入表头
        _=sheet.cell(row=1,column=field,value=str(fields[field-1]))

    row1=1
    col1=0
    for row1 in range(2,len(data)+2):  # 写入数据
        for col1 in range(1,len(data[row1-2])+1):
            _=sheet.cell(row=row1,column=col1,value=str(data[row1-2][col1-1]))

    wb.save(filename=wbname)
    print("保存成功")